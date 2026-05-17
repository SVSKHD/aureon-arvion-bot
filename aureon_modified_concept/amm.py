# aureon_modified_mock.py
#
# Paper-trade MOCK of the Aureon Modified XAUUSD anchor-breakout strategy.
#
# This file is the EDIT TARGET — it gets the following upgrades from the
# original:
#   1. EOD close moved to 23:00 broker time (was 23:59).
#   2. Bid/ask spread is now modelled honestly. Entries fill at the side
#      a real stop order would fill at (ask for long, bid for short); exits
#      fill at the side a real close would fill at (bid for long, ask for
#      short). The actual bid/ask at the moment the trigger fires is the
#      fill price — NOT the theoretical trigger/lock level. This exposes
#      slippage on fast moves the way live trading would.
#   3. The flat COST_PER_TRADE becomes COMMISSION_PER_TRADE — spread is
#      now captured naturally through bid/ask fills, so this constant only
#      needs to cover commission and residual slippage.
#   4. Every print message that announces an event shows both broker time
#      and IST so the operator can correlate on either clock.
#   5. Trade log schema adds entry_bid/entry_ask/exit_bid/exit_ask/spread
#      fields for full audit.
#
# All other behaviour matches the original:
#   - Connects to MT5 for live tick data. PLACES NO REAL ORDERS.
#   - At 02:00 broker time each day, anchors to the open of the first M1 bar
#     and registers two INDEPENDENT virtual stop orders (no OCO link):
#         buy  stop at anchor + TRIGGER_DISTANCE
#         sell stop at anchor - TRIGGER_DISTANCE
#   - Each leg runs its own state machine (pending -> active -> closed).
#     Both can fire the same day; the second is NOT cancelled when the first
#     fills. This is the modified-vs-base behaviour.
#   - Per-leg trail lock arms at +TRAIL_ACTIVATION profit, then trails behind
#     the extreme by TRAIL_DISTANCE. No TP cap. Hard SL at entry +/- HARD_SL,
#     bypassed once the lock arms.
#   - At 23:00 broker time any still-active leg is closed at the current
#     bid (long) or ask (short); any still-pending leg is cancelled. Daily
#     summary prints at the 02:00 rollover or on Ctrl+C.

import csv
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

import MetaTrader5 as mt5

# Optional: openpyxl for the daily XLSX summary. If it's not installed the
# bot still runs — XLSX just gets skipped and a one-time warning prints.
# Install with: pip install openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# ============================================================================
# CONFIG
# ============================================================================

SYMBOL = "XAUUSD"
TIMEFRAME = mt5.TIMEFRAME_M1

# Anchor: open of the M1 bar at 02:00 broker time.
ANCHOR_HOUR = 2
ANCHOR_MINUTE = 0

# EOD: 23:00 broker time. Any still-active leg is closed at the current
# realistic side (bid for long, ask for short); any still-pending leg is
# cancelled. The 1-hour gap before the 02:00 anchor of the next day acts
# as a buffer for late ticks / restarts / broker hiccups.
EOD_HOUR = 23
EOD_MINUTE = 0

# Strategy parameters (must match the backtest that produced the validated
# OOS curve). Changing any of these invalidates the OOS validation.
TRIGGER_DISTANCE = 5.00     # USD from anchor — places buy/sell stops at +/-$5
HARD_SL          = 5.00     # USD from entry — disconnect parachute
TRAIL_DISTANCE   = 0.30     # USD behind extreme — trail lock
TRAIL_ACTIVATION = 0.30     # USD profit required before the lock arms

# Cost components — separated from spread now that spread is modelled directly
# through bid/ask fills.
#   COMMISSION_PER_TRADE: USD/oz round-trip. Broker commission plus residual
#     slippage you'd see live (order routing latency, partial fills, etc.)
#     not captured by the bid/ask model. Typical retail XAUUSD: $0.03-0.07/oz.
#   FIXED_SPREAD_FALLBACK: USD/oz. Used ONLY when MT5's tick.ask is unavailable
#     (some demo accounts return ask==bid). Lets the bot keep simulating spread
#     even when the broker's quote feed is degenerate.
COMMISSION_PER_TRADE = 0.05
FIXED_SPREAD_FALLBACK = 0.20

# Position sizing — adjust to match the funded account profile this mock
# runs alongside. The mock doesn't trade, but the dollar-P&L column in the
# trade log uses this so the numbers line up with the live account.
LOT_SIZE = 0.20             # per leg; 0.33 for 25K, 0.50 for 50K, 0.90 for 100K

# Lot tiers shown side-by-side in the daily XLSX so you can see what the same
# day's P&L would have been at each funded-account size. The dollar P&L scales
# linearly with lot size (same trade decisions, just bigger position), so this
# is a clean what-if view. Keys become column labels in the spreadsheet.
LOT_SIZE_TIERS = {
    "current": LOT_SIZE,
    "0.20":   0.20,
    "0.33 (25K)":  0.33,
    "0.50 (50K)":  0.50,
    "0.90 (100K)": 0.90,
}

# === Account balance simulation ===
# Starting capital is used the FIRST time the bot runs (or whenever the
# balance file is missing). Once a balance file exists it's the source of
# truth — edit STARTING_CAPITAL freely; it only takes effect on a fresh
# initialization. To reset the bot to STARTING_CAPITAL, delete the balance
# file: {SYMBOL_PREFIX}_aureon_mod_balance.json
STARTING_CAPITAL = 48500.00

POLL_SECONDS = 1            # tick poll cadence
RETRY_SECONDS = 5           # back-off after a transient MT5 error
HEARTBEAT_WIDTH = 220       # padding for the live one-liner

# Replay M1 bars from 02:00 -> now on startup so a restart doesn't lose state.
# Synthesises any leg fills / exits that occurred while the mock was offline,
# tagging them backfill=True in the trade log.
# NOTE: backfill uses bar close as a proxy for bid and applies the fixed
# spread fallback. It will NOT reproduce the precise bid/ask realism of the
# live tick stream. Acceptable for catch-up of short outages; trades are
# tagged so you can separate them from live-stream trades.
BACKFILL_ON_START = True

IST = ZoneInfo("Asia/Kolkata")

# Output files — prefixed so this mock and the existing tracker can run
# side-by-side without colliding on writes.
SYMBOL_PREFIX = SYMBOL.lower()
LOG_PATH       = f"{SYMBOL_PREFIX}_aureon_mod_daily.csv"
TRADE_LOG_PATH = f"{SYMBOL_PREFIX}_aureon_mod_trades.csv"
EVENT_LOG_PATH = f"{SYMBOL_PREFIX}_aureon_mod_events.jsonl"
BALANCE_PATH   = f"{SYMBOL_PREFIX}_aureon_mod_balance.json"
XLSX_PATH      = f"{SYMBOL_PREFIX}_aureon_mod_daily.xlsx"

DAILY_LOG_FIELDS = [
    "date",
    "anchor_price",
    "anchor_server_time",
    "day_start_balance",
    "buy_stop_price",
    "sell_stop_price",
    "long_status",
    "long_entry_price",
    "long_exit_price",
    "long_pnl_oz",
    "long_pnl_usd",
    "long_reason",
    "short_status",
    "short_entry_price",
    "short_exit_price",
    "short_pnl_oz",
    "short_pnl_usd",
    "short_reason",
    "day_pnl_oz",
    "day_pnl_usd",
    "day_end_balance",
    "cumulative_pnl_usd",
    "last_price",
    "last_update_server",
    "last_update_ist",
]

# Trade log schema — adds bid/ask audit fields so every closed leg can be
# fully reconstructed (exact spread at entry and exit, which side was used).
TRADE_LOG_FIELDS = [
    "trade_id", "session_date",
    "leg", "lot", "backfill",
    "anchor_price", "anchor_server_time",
    "trigger_price",
    "entry_price", "entry_bid", "entry_ask", "spread_at_entry",
    "entry_server_time", "entry_ist_time",
    "extreme_seen",
    "lock_armed", "lock_price_final",
    "exit_price", "exit_bid", "exit_ask", "spread_at_exit",
    "exit_server_time", "exit_ist_time", "exit_reason",
    "duration_seconds",
    "mfe_price", "mae_price",
    "gross_pnl_oz", "commission", "net_pnl_oz", "net_pnl_usd",
    "opposite_leg_fired",
]


# ============================================================================
# MT5 HELPERS  (same idiom as the live tracker)
# ============================================================================

def connect_mt5():
    if not mt5.initialize():
        raise RuntimeError(f"MT5 initialize failed: {mt5.last_error()}")
    if not mt5.symbol_select(SYMBOL, True):
        raise RuntimeError(f"Symbol select failed: {SYMBOL}")


def get_ist_now():
    return datetime.now(IST)


def broker_dt(epoch: int) -> datetime:
    """MT5 epoch -> naive datetime matching broker wall-clock. TZ-agnostic."""
    return datetime.fromtimestamp(epoch, tz=timezone.utc).replace(tzinfo=None)


def to_mt5_dt(dt: datetime) -> datetime:
    """Tag a naive broker datetime as UTC before passing to MT5. The wrapper
    converts naive datetimes using local TZ, which silently breaks lookups
    on any machine that isn't on broker time."""
    return dt.replace(tzinfo=timezone.utc)


def get_server_now(symbol: str) -> datetime:
    tick = mt5.symbol_info_tick(symbol)
    if tick is None:
        raise RuntimeError(f"No tick data for {symbol}")
    return broker_dt(tick.time)


def get_bid_ask(symbol: str) -> tuple:
    """Return (bid, ask) from the current tick. Falls back to (bid, bid+spread)
    if the broker's quote feed reports ask <= bid (some demo accounts do this).
    Returns (None, None) only if MT5 has no tick data at all."""
    tick = mt5.symbol_info_tick(symbol)
    if tick is None:
        return None, None
    bid = float(tick.bid)
    ask = float(tick.ask) if tick.ask else 0.0
    if ask <= bid:
        ask = bid + FIXED_SPREAD_FALLBACK
    return bid, ask


def get_current_anchor_time(server_now: datetime) -> datetime:
    """Most recent 02:00 anchor in broker time. Steps back across weekends."""
    anchor = server_now.replace(hour=ANCHOR_HOUR, minute=ANCHOR_MINUTE,
                                 second=0, microsecond=0)
    if server_now < anchor:
        anchor -= timedelta(days=1)
    while anchor.weekday() >= 5:  # Sat/Sun
        anchor -= timedelta(days=1)
    return anchor


def get_pip_size(symbol: str) -> float:
    info = mt5.symbol_info(symbol)
    if info is None:
        raise RuntimeError(f"Symbol info not found: {symbol}")
    if info.digits in (3, 5):
        return info.point * 10
    return info.point


def fetch_anchor(symbol: str, anchor_time: datetime):
    """Fetch the 02:00 anchor M1 candle. Verifies timestamp match — catches
    data gaps where the requested bar doesn't exist (e.g. broker downtime)."""
    rates = mt5.copy_rates_from(symbol, TIMEFRAME, to_mt5_dt(anchor_time), 1)
    if rates is None or len(rates) == 0:
        raise RuntimeError(f"Anchor candle not found at {anchor_time}")
    candle = rates[0]
    candle_dt = broker_dt(candle["time"])
    if candle_dt != anchor_time:
        raise RuntimeError(
            f"Anchor candle mismatch: wanted {anchor_time}, got {candle_dt}"
        )
    return {"price": float(candle["open"]), "server_time": candle_dt}


def order_calc_pnl(side: str, symbol: str, lot: float,
                   entry: float, exit_price: float):
    """Broker-aware account-currency P&L. Returns None if MT5 unavailable so
    the caller can fall back to manual contract-size math."""
    if not hasattr(mt5, "order_calc_profit"):
        return None
    action = mt5.ORDER_TYPE_BUY if side == "long" else mt5.ORDER_TYPE_SELL
    try:
        result = mt5.order_calc_profit(action, symbol, lot, entry, exit_price)
        return None if result is None else float(result)
    except Exception:
        return None


def fmt_times(server_dt: datetime, ist_dt: datetime) -> str:
    """Compact '[broker HH:MM:SS  IST HH:MM:SS]' for console messages.
    Both clocks shown together so the operator can correlate either way."""
    return f"[broker {server_dt:%H:%M:%S}  IST {ist_dt:%H:%M:%S}]"


def get_mt5_balance() -> float:
    """Read the current MT5 account balance (cash, not equity). Used as the
    day_start_balance at each session_start so the paper bot tracks against
    the real account. Returns None if account_info isn't available — caller
    falls back to the persisted paper balance in that case."""
    if not hasattr(mt5, "account_info"):
        return None
    try:
        info = mt5.account_info()
        if info is None:
            return None
        return float(info.balance)
    except Exception:
        return None


def get_mt5_equity() -> float:
    """Account equity = balance + floating P&L from open positions. Useful for
    the daily report so the operator can compare paper P&L against the live
    account's realised+unrealised total. None if MT5 can't supply it."""
    if not hasattr(mt5, "account_info"):
        return None
    try:
        info = mt5.account_info()
        if info is None:
            return None
        return float(info.equity)
    except Exception:
        return None


# ============================================================================
# LEG STATE
# ============================================================================

def make_leg(side: str, trigger_price: float, anchor_price: float) -> dict:
    """Construct a fresh leg state. side = 'long' | 'short'."""
    if side == "long":
        hard_sl_price = trigger_price - HARD_SL
    else:
        hard_sl_price = trigger_price + HARD_SL
    return {
        "side": side,
        "status": "pending",          # pending | active | closed | cancelled
        "trigger_price": round(trigger_price, 2),
        "hard_sl_price": round(hard_sl_price, 2),
        "anchor_price": round(anchor_price, 2),
        "trade_id": None,
        "entry_price": None,
        "entry_bid": None,            # bid at the moment of trigger
        "entry_ask": None,            # ask at the moment of trigger
        "spread_at_entry": None,
        "entry_server_time": None,
        "entry_ist_time": None,
        "extreme_price": None,        # peak bid (long) or trough ask (short)
        "lock_armed": False,
        "lock_price": None,
        "exit_price": None,
        "exit_bid": None,
        "exit_ask": None,
        "spread_at_exit": None,
        "exit_server_time": None,
        "exit_ist_time": None,
        "exit_reason": None,
        "mfe_price": 0.0,             # max favourable excursion in price terms
        "mae_price": 0.0,             # max adverse excursion in price terms
        "backfill": False,
    }


def trigger_hit(leg: dict, bid: float, ask: float) -> bool:
    """Realistic stop-order trigger:
        BUY stop  fires when ASK reaches the trigger (cross the offer to buy).
        SELL stop fires when BID reaches the trigger (cross the bid to sell).
    Checking against the wrong side under-fires longs and over-fires shorts."""
    if leg["side"] == "long":
        return ask >= leg["trigger_price"]
    return bid <= leg["trigger_price"]


def hard_sl_hit(leg: dict, bid: float, ask: float) -> bool:
    """Hard SL — only checked while lock NOT armed.
        LONG SL: triggers when BID drops to or below SL (sell at bid).
        SHORT SL: triggers when ASK rises to or above SL (buy at ask)."""
    if leg["side"] == "long":
        return bid <= leg["hard_sl_price"]
    return ask >= leg["hard_sl_price"]


def lock_hit(leg: dict, bid: float, ask: float) -> bool:
    """Lock breach — only fires once lock is armed.
        LONG lock: triggers when BID drops to or below the lock.
        SHORT lock: triggers when ASK rises to or above the lock."""
    if not leg["lock_armed"] or leg["lock_price"] is None:
        return False
    if leg["side"] == "long":
        return bid <= leg["lock_price"]
    return ask >= leg["lock_price"]


def update_extreme_and_lock(leg: dict, bid: float, ask: float):
    """Update favourable extreme and (re)position the trail lock.
    LONG: extreme is the peak BID seen (highest sellable price).
    SHORT: extreme is the trough ASK seen (lowest buyback price).
    Returns (lock_was_armed_this_tick, lock_advanced_this_tick) for logging."""
    armed_now = False
    advanced_now = False

    if leg["side"] == "long":
        # Long extreme tracked on BID — the price you could sell at.
        if bid > leg["extreme_price"]:
            leg["extreme_price"] = bid
            favourable = bid - leg["entry_price"]
            if favourable > leg["mfe_price"]:
                leg["mfe_price"] = round(favourable, 4)
        else:
            adverse = leg["entry_price"] - bid
            if adverse > leg["mae_price"]:
                leg["mae_price"] = round(max(0.0, adverse), 4)

        profit_so_far = leg["extreme_price"] - leg["entry_price"]
        if not leg["lock_armed"]:
            if profit_so_far >= TRAIL_ACTIVATION:
                leg["lock_armed"] = True
                leg["lock_price"] = round(leg["extreme_price"] - TRAIL_DISTANCE, 2)
                armed_now = True
        else:
            new_lock = round(leg["extreme_price"] - TRAIL_DISTANCE, 2)
            if new_lock > leg["lock_price"]:
                leg["lock_price"] = new_lock
                advanced_now = True
    else:
        # Short extreme tracked on ASK — the price you could buy back at.
        if ask < leg["extreme_price"]:
            leg["extreme_price"] = ask
            favourable = leg["entry_price"] - ask
            if favourable > leg["mfe_price"]:
                leg["mfe_price"] = round(favourable, 4)
        else:
            adverse = ask - leg["entry_price"]
            if adverse > leg["mae_price"]:
                leg["mae_price"] = round(max(0.0, adverse), 4)

        profit_so_far = leg["entry_price"] - leg["extreme_price"]
        if not leg["lock_armed"]:
            if profit_so_far >= TRAIL_ACTIVATION:
                leg["lock_armed"] = True
                leg["lock_price"] = round(leg["extreme_price"] + TRAIL_DISTANCE, 2)
                armed_now = True
        else:
            new_lock = round(leg["extreme_price"] + TRAIL_DISTANCE, 2)
            if new_lock < leg["lock_price"]:
                leg["lock_price"] = new_lock
                advanced_now = True

    return armed_now, advanced_now


def close_leg(leg: dict, exit_bid: float, exit_ask: float, reason: str,
              server_now: datetime, ist_now: datetime,
              contract_size: float, opposite_fired: bool):
    """Finalize a leg at REALISTIC fill prices:
        LONG exit fills at the BID (selling to close).
        SHORT exit fills at the ASK (buying to close).
    Slippage past the theoretical lock/SL level (on fast moves) is captured
    automatically — the fill price is whatever bid/ask is at the moment we
    detect the trigger, not the trigger level itself."""
    if reason == "hard_sl":
        assert not leg["lock_armed"], (
            f"INVARIANT VIOLATION: hard SL fired after lock armed on {leg['side']}"
        )

    leg["status"] = "closed"
    leg["exit_bid"] = round(exit_bid, 2)
    leg["exit_ask"] = round(exit_ask, 2)
    leg["spread_at_exit"] = round(exit_ask - exit_bid, 2)
    # Exit fill: bid for long (sell to close), ask for short (buy to close).
    fill = exit_bid if leg["side"] == "long" else exit_ask
    leg["exit_price"] = round(fill, 2)
    leg["exit_server_time"] = server_now
    leg["exit_ist_time"] = ist_now
    leg["exit_reason"] = reason

    if leg["side"] == "long":
        gross = leg["exit_price"] - leg["entry_price"]
    else:
        gross = leg["entry_price"] - leg["exit_price"]

    # gross_pnl_oz already reflects the bid/ask spread cost (LONG bought at
    # ask, sold at bid; SHORT sold at bid, bought at ask). commission is the
    # only remaining round-trip cost to subtract.
    leg["gross_pnl_oz"] = round(gross, 2)
    leg["commission"] = COMMISSION_PER_TRADE
    leg["net_pnl_oz"] = round(gross - COMMISSION_PER_TRADE, 2)

    broker_pnl = order_calc_pnl(leg["side"], SYMBOL, LOT_SIZE,
                                leg["entry_price"], leg["exit_price"])
    if broker_pnl is None:
        broker_pnl = gross * LOT_SIZE * contract_size
    commission_usd = COMMISSION_PER_TRADE * LOT_SIZE * contract_size
    leg["net_pnl_usd"] = round(broker_pnl - commission_usd, 2)

    try:
        leg["duration_seconds"] = round(
            (leg["exit_server_time"] - leg["entry_server_time"]).total_seconds(), 1
        )
    except Exception:
        leg["duration_seconds"] = 0.0

    leg["opposite_leg_fired"] = opposite_fired


# ============================================================================
# EVENT / TRADE / DAILY LOGGING
# ============================================================================

def log_event(event_type: str, data: dict,
              server_now: datetime, ist_now: datetime):
    """Append one JSON line to the event log. Soft-fails so a disk hiccup
    can't kill the tracker."""
    event = {
        "ts_server": server_now.isoformat(sep=" "),
        "ts_ist": ist_now.isoformat(sep=" "),
        "event": event_type,
        "data": data,
    }
    try:
        with open(EVENT_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(event, default=str) + "\n")
    except Exception as e:
        print(f"[event log write failed: {e}]")


def load_recent_events(path: str, n: int = 100) -> list:
    if not os.path.exists(path):
        return []
    events = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                events.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return events[-n:]


def verify_continuity(events: list) -> str:
    """Quick health check on the event log — printed in the startup banner."""
    if not events:
        return "first run — no prior event history"
    finalized = [e for e in events if e["event"] == "day_finalized"]
    if not finalized:
        return f"no day_finalized events yet ({len(events)} prior events)"
    last = finalized[-1]
    prev = last["data"].get("session_date", "?")
    pnl = last["data"].get("day_pnl_oz", 0)
    return f"last finalized: {prev}  (day P&L ${pnl:+.2f}/oz)"


def ensure_trade_log_header(path: str):
    if not os.path.exists(path):
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=TRADE_LOG_FIELDS)
            writer.writeheader()


def append_trade_log(leg: dict, session_date: str):
    """One row per closed leg-trade."""
    row = {
        "trade_id": leg.get("trade_id"),
        "session_date": session_date,
        "leg": leg["side"],
        "lot": LOT_SIZE,
        "backfill": leg.get("backfill", False),
        "anchor_price": leg["anchor_price"],
        "anchor_server_time": "",
        "trigger_price": leg["trigger_price"],
        "entry_price": leg["entry_price"],
        "entry_bid": leg.get("entry_bid"),
        "entry_ask": leg.get("entry_ask"),
        "spread_at_entry": leg.get("spread_at_entry"),
        "entry_server_time": leg["entry_server_time"],
        "entry_ist_time": leg["entry_ist_time"],
        "extreme_seen": leg.get("extreme_price"),
        "lock_armed": leg["lock_armed"],
        "lock_price_final": leg.get("lock_price"),
        "exit_price": leg["exit_price"],
        "exit_bid": leg.get("exit_bid"),
        "exit_ask": leg.get("exit_ask"),
        "spread_at_exit": leg.get("spread_at_exit"),
        "exit_server_time": leg["exit_server_time"],
        "exit_ist_time": leg["exit_ist_time"],
        "exit_reason": leg["exit_reason"],
        "duration_seconds": leg.get("duration_seconds", 0),
        "mfe_price": leg.get("mfe_price", 0),
        "mae_price": leg.get("mae_price", 0),
        "gross_pnl_oz": leg.get("gross_pnl_oz", 0),
        "commission": leg.get("commission", COMMISSION_PER_TRADE),
        "net_pnl_oz": leg.get("net_pnl_oz", 0),
        "net_pnl_usd": leg.get("net_pnl_usd", 0),
        "opposite_leg_fired": leg.get("opposite_leg_fired", False),
    }
    for k in ("entry_server_time", "entry_ist_time",
              "exit_server_time", "exit_ist_time"):
        if isinstance(row[k], datetime):
            row[k] = row[k].isoformat(sep=" ")

    with open(TRADE_LOG_PATH, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TRADE_LOG_FIELDS)
        writer.writerow(row)


def load_balance() -> dict:
    """Load the running balance state, or initialize from STARTING_CAPITAL
    if the balance file doesn't exist yet. The file is the source of truth
    once it exists — changing STARTING_CAPITAL has no effect until the file
    is deleted."""
    if not os.path.exists(BALANCE_PATH):
        return {
            "starting_capital": STARTING_CAPITAL,
            "current_balance": STARTING_CAPITAL,
            "trade_count": 0,
            "last_updated": None,
        }
    try:
        with open(BALANCE_PATH, "r") as f:
            state = json.load(f)
        # Defensive: backfill any missing fields if file is from an older run.
        state.setdefault("starting_capital", STARTING_CAPITAL)
        state.setdefault("current_balance", state["starting_capital"])
        state.setdefault("trade_count", 0)
        return state
    except (json.JSONDecodeError, OSError):
        # Corrupt or unreadable — fall back to fresh init but DON'T overwrite
        # the file. Operator should investigate before next save.
        return {
            "starting_capital": STARTING_CAPITAL,
            "current_balance": STARTING_CAPITAL,
            "trade_count": 0,
            "last_updated": None,
        }


def save_balance(state: dict):
    """Atomic write of the balance file. Same .tmp+rename idiom as the
    daily CSV — survives a kill mid-write."""
    state["last_updated"] = datetime.utcnow().isoformat(sep=" ")
    tmp = BALANCE_PATH + ".tmp"
    try:
        with open(tmp, "w") as f:
            json.dump(state, f, indent=2)
        try:
            os.replace(tmp, BALANCE_PATH)
        except PermissionError:
            time.sleep(0.2)
            os.replace(tmp, BALANCE_PATH)
    except Exception as e:
        print(f"[balance save failed: {e}]")


def book_trade(leg: dict, balance_state: dict):
    """Apply a closed leg's net_pnl_usd to the running balance and persist.
    Called immediately after a leg is closed AND its CSV row is written, so
    the balance file and the trade log are always consistent."""
    delta = float(leg.get("net_pnl_usd", 0.0) or 0.0)
    balance_state["current_balance"] = round(balance_state["current_balance"] + delta, 2)
    balance_state["trade_count"] = balance_state.get("trade_count", 0) + 1
    save_balance(balance_state)


def load_daily_log(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    records = {}
    with open(path, "r", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records[row["date"]] = row
    return records


def save_daily_log(path: str, records: dict):
    """Atomic write — .tmp + rename. Prevents partial writes on crash."""
    tmp = path + ".tmp"
    with open(tmp, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=DAILY_LOG_FIELDS)
        writer.writeheader()
        for date in sorted(records.keys()):
            writer.writerow(records[date])
    try:
        os.replace(tmp, path)
    except PermissionError:
        # Windows can transiently lock the file (e.g. when Excel has it
        # open). Sleep briefly and retry once; let the caller decide what
        # to do if the second attempt also fails.
        time.sleep(0.2)
        os.replace(tmp, path)


def compute_lot_pnl_variants(day_pnl_oz: float, contract_size: float) -> dict:
    """Map each LOT_SIZE_TIERS entry to its dollar P&L at that lot size.
    Linear scaling: pnl_at_lot = day_pnl_oz * lot * contract_size. Assumes the
    same trade decisions at different lot sizes (true here — strategy doesn't
    depend on lot size, only position sizing does)."""
    out = {}
    for label, lot in LOT_SIZE_TIERS.items():
        out[label] = round(day_pnl_oz * lot * contract_size, 2)
    return out


# Static schema for the daily XLSX. Columns are arranged left-to-right in
# operator-readable order: when/anchor → both legs' results → day total
# scaled across funded tiers → balance trajectory.
XLSX_HEADERS = [
    "Date", "Anchor", "Buy Stop", "Sell Stop",
    "Legs Filled",
    "Long Status", "Long Entry", "Long Exit", "Long Reason", "Long P&L ($/oz)",
    "Short Status", "Short Entry", "Short Exit", "Short Reason", "Short P&L ($/oz)",
    "Day P&L ($/oz)",
    "P&L @ current", "P&L @ 0.20", "P&L @ 0.33 (25K)", "P&L @ 0.50 (50K)", "P&L @ 0.90 (100K)",
    "Day Start Balance", "Day End Balance", "Cumulative P&L", "Cumulative %",
    "MT5 Balance (start)", "MT5 Equity (close)",
    "Notes",
]


def update_daily_xlsx(daily_rows: list, balance_state: dict, contract_size: float):
    """Rewrite the daily XLSX from the running list of finalized day rows.
    Called once per day at EOD / day-rollover / shutdown (NOT every poll —
    XLSX rewrites the whole file each save, so per-poll would be wasteful).

    daily_rows is a list of dicts, one per trading day, each with the keys
    expected by _row_for_xlsx() below. Caller maintains the list in
    chronological order and persists it via the regular daily CSV; this XLSX
    is a parallel pretty-print of the same data."""
    if not XLSX_AVAILABLE:
        # Print once per session if missing — don't spam every save
        if not getattr(update_daily_xlsx, "_warned", False):
            print(f"[xlsx skipped — install openpyxl to enable: pip install openpyxl]")
            update_daily_xlsx._warned = True
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Aureon Daily"

        # Header row — bold, light fill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for col_idx, label in enumerate(XLSX_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(daily_rows, start=2):
            for c_idx, key in enumerate(XLSX_HEADERS, start=1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))

        # Auto-size columns (cheap heuristic — max content width per col)
        for col_idx, label in enumerate(XLSX_HEADERS, start=1):
            max_len = len(label)
            for row in daily_rows:
                v = row.get(label, "")
                vs = str(v) if v is not None else ""
                if len(vs) > max_len:
                    max_len = len(vs)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

        # Atomic write (.tmp + rename) so an Excel-open-on-the-file Windows
        # PermissionError doesn't truncate the workbook to zero.
        tmp = XLSX_PATH + ".tmp"
        wb.save(tmp)
        try:
            os.replace(tmp, XLSX_PATH)
        except PermissionError:
            time.sleep(0.3)
            os.replace(tmp, XLSX_PATH)
    except Exception as e:
        print(f"[xlsx write failed: {e}]")


def daily_xlsx_row(active_anchor: datetime, anchor_price: float,
                   buy_stop: float, sell_stop: float,
                   long_leg: dict, short_leg: dict,
                   day_start_balance: float, balance_state: dict,
                   contract_size: float,
                   mt5_balance_at_start: float = None,
                   mt5_equity_at_close: float = None,
                   notes: str = "") -> dict:
    """Build one Excel row dict for a finalized day. Keys must match XLSX_HEADERS
    exactly so update_daily_xlsx() can drop the value into the right column."""
    def s(leg, key, default=""):
        v = leg.get(key)
        return v if v is not None else default

    day_oz = (long_leg.get("net_pnl_oz", 0) or 0) + (short_leg.get("net_pnl_oz", 0) or 0)
    variants = compute_lot_pnl_variants(day_oz, contract_size)
    end_bal = balance_state["current_balance"]
    cum = round(end_bal - balance_state["starting_capital"], 2)
    cum_pct = (cum / balance_state["starting_capital"] * 100.0
               if balance_state["starting_capital"] else 0.0)
    n_filled = sum(1 for l in (long_leg, short_leg) if l["status"] == "closed")

    return {
        "Date": active_anchor.date().isoformat(),
        "Anchor": round(anchor_price, 2),
        "Buy Stop": round(buy_stop, 2),
        "Sell Stop": round(sell_stop, 2),
        "Legs Filled": n_filled,
        "Long Status": long_leg["status"],
        "Long Entry": s(long_leg, "entry_price"),
        "Long Exit": s(long_leg, "exit_price"),
        "Long Reason": s(long_leg, "exit_reason"),
        "Long P&L ($/oz)": s(long_leg, "net_pnl_oz"),
        "Short Status": short_leg["status"],
        "Short Entry": s(short_leg, "entry_price"),
        "Short Exit": s(short_leg, "exit_price"),
        "Short Reason": s(short_leg, "exit_reason"),
        "Short P&L ($/oz)": s(short_leg, "net_pnl_oz"),
        "Day P&L ($/oz)": round(day_oz, 2),
        "P&L @ current": variants["current"],
        "P&L @ 0.20": variants["0.20"],
        "P&L @ 0.33 (25K)": variants["0.33 (25K)"],
        "P&L @ 0.50 (50K)": variants["0.50 (50K)"],
        "P&L @ 0.90 (100K)": variants["0.90 (100K)"],
        "Day Start Balance": round(day_start_balance, 2),
        "Day End Balance": round(end_bal, 2),
        "Cumulative P&L": cum,
        "Cumulative %": round(cum_pct, 3),
        "MT5 Balance (start)": (round(mt5_balance_at_start, 2)
                                if mt5_balance_at_start is not None else ""),
        "MT5 Equity (close)": (round(mt5_equity_at_close, 2)
                               if mt5_equity_at_close is not None else ""),
        "Notes": notes,
    }


def load_xlsx_rows_from_csv() -> list:
    """On startup, rebuild the in-memory daily_rows list from the persisted
    CSV daily log + balance file. Lets the XLSX preserve history across
    restarts without keeping a separate state file. Falls back to empty list
    if the CSV doesn't exist yet."""
    if not os.path.exists(LOG_PATH):
        return []
    rows = []
    with open(LOG_PATH, "r", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            # The CSV has fewer columns than XLSX_HEADERS; just remap what
            # we can and leave the rest blank. Multi-lot recompute happens
            # on the fly so the same CSV history can produce a fresh XLSX.
            try:
                day_oz = float(r.get("day_pnl_oz") or 0)
            except ValueError:
                day_oz = 0.0
            variants = compute_lot_pnl_variants(day_oz, 100.0)
            rows.append({
                "Date": r.get("date", ""),
                "Anchor": r.get("anchor_price", ""),
                "Buy Stop": r.get("buy_stop_price", ""),
                "Sell Stop": r.get("sell_stop_price", ""),
                "Legs Filled": "",
                "Long Status": r.get("long_status", ""),
                "Long Entry": r.get("long_entry_price", ""),
                "Long Exit": r.get("long_exit_price", ""),
                "Long Reason": r.get("long_reason", ""),
                "Long P&L ($/oz)": r.get("long_pnl_oz", ""),
                "Short Status": r.get("short_status", ""),
                "Short Entry": r.get("short_entry_price", ""),
                "Short Exit": r.get("short_exit_price", ""),
                "Short Reason": r.get("short_reason", ""),
                "Short P&L ($/oz)": r.get("short_pnl_oz", ""),
                "Day P&L ($/oz)": day_oz,
                "P&L @ current": variants["current"],
                "P&L @ 0.20": variants["0.20"],
                "P&L @ 0.33 (25K)": variants["0.33 (25K)"],
                "P&L @ 0.50 (50K)": variants["0.50 (50K)"],
                "P&L @ 0.90 (100K)": variants["0.90 (100K)"],
                "Day Start Balance": r.get("day_start_balance", ""),
                "Day End Balance": r.get("day_end_balance", ""),
                "Cumulative P&L": r.get("cumulative_pnl_usd", ""),
                "Cumulative %": "",
                "MT5 Balance (start)": "",
                "MT5 Equity (close)": "",
                "Notes": "",
            })
    return rows


def build_daily_record(active_anchor: datetime, anchor_price: float,
                       buy_stop: float, sell_stop: float,
                       long_leg: dict, short_leg: dict,
                       last_price: float,
                       server_now: datetime, ist_now: datetime,
                       day_start_balance: float,
                       balance_state: dict) -> dict:
    """Flatten current day's state into one CSV row. Called every poll so the
    daily file always reflects the latest known state for this session.
    The balance columns (day_start_balance, day_end_balance, cumulative_pnl_usd)
    are updated live so the file is always a current account statement."""
    def leg_cols(leg: dict) -> dict:
        return {
            "status": leg["status"],
            "entry": leg.get("entry_price") or "",
            "exit": leg.get("exit_price") or "",
            "pnl_oz": leg.get("net_pnl_oz") or "",
            "pnl_usd": leg.get("net_pnl_usd") or "",
            "reason": leg.get("exit_reason") or "",
        }
    long_c = leg_cols(long_leg)
    short_c = leg_cols(short_leg)
    day_oz = (long_leg.get("net_pnl_oz", 0) or 0) + (short_leg.get("net_pnl_oz", 0) or 0)
    day_usd = (long_leg.get("net_pnl_usd", 0) or 0) + (short_leg.get("net_pnl_usd", 0) or 0)
    current_bal = balance_state["current_balance"]
    cum_pnl = round(current_bal - balance_state["starting_capital"], 2)

    return {
        "date": active_anchor.date().isoformat(),
        "anchor_price": round(anchor_price, 2),
        "anchor_server_time": active_anchor.isoformat(sep=" "),
        "day_start_balance": round(day_start_balance, 2),
        "buy_stop_price": round(buy_stop, 2),
        "sell_stop_price": round(sell_stop, 2),
        "long_status": long_c["status"],
        "long_entry_price": long_c["entry"],
        "long_exit_price": long_c["exit"],
        "long_pnl_oz": long_c["pnl_oz"],
        "long_pnl_usd": long_c["pnl_usd"],
        "long_reason": long_c["reason"],
        "short_status": short_c["status"],
        "short_entry_price": short_c["entry"],
        "short_exit_price": short_c["exit"],
        "short_pnl_oz": short_c["pnl_oz"],
        "short_pnl_usd": short_c["pnl_usd"],
        "short_reason": short_c["reason"],
        "day_pnl_oz": round(day_oz, 2),
        "day_pnl_usd": round(day_usd, 2),
        "day_end_balance": round(current_bal, 2),
        "cumulative_pnl_usd": cum_pnl,
        "last_price": round(last_price, 2),
        "last_update_server": server_now.isoformat(sep=" "),
        "last_update_ist": ist_now.isoformat(sep=" "),
    }


# ============================================================================
# DAILY SUMMARY
# ============================================================================

def print_daily_summary(session_date: str, long_leg: dict, short_leg: dict,
                        contract_size: float,
                        day_start_balance: float = None,
                        balance_state: dict = None):
    """Compact end-of-day report — printed at 23:00 EOD, 02:00 rollover, Ctrl+C.
    If balance state is supplied, also prints the day's balance trajectory
    (starting balance, day P&L, ending balance, lifetime cumulative)."""
    def fmt_leg(leg: dict) -> str:
        s = leg["status"]
        if s == "pending":
            return "PENDING (never triggered)"
        if s == "cancelled":
            return "CANCELLED at EOD (never triggered)"
        return (
            f"{leg['entry_price']:.2f} -> {leg['exit_price']:.2f}  "
            f"({leg['exit_reason']})  "
            f"gross ${leg.get('gross_pnl_oz', 0):+.2f}/oz  "
            f"net ${leg.get('net_pnl_oz', 0):+.2f}/oz  "
            f"(${leg.get('net_pnl_usd', 0):+.2f})"
        )

    day_oz = (long_leg.get("net_pnl_oz", 0) or 0) + (short_leg.get("net_pnl_oz", 0) or 0)
    day_usd = (long_leg.get("net_pnl_usd", 0) or 0) + (short_leg.get("net_pnl_usd", 0) or 0)
    n_filled = sum(1 for l in (long_leg, short_leg) if l["status"] == "closed")

    print()
    print("=" * 72)
    print(f"  AUREON MODIFIED — DAILY SUMMARY  ·  {session_date}  ·  {SYMBOL}")
    print("=" * 72)
    if day_start_balance is not None:
        print(f"  Starting balance   : ${day_start_balance:>12,.2f}")
    print(f"  Anchor             : ${long_leg['anchor_price']:.2f}")
    print(f"  Buy stop (long)    : ${long_leg['trigger_price']:.2f}")
    print(f"  Sell stop (short)  : ${short_leg['trigger_price']:.2f}")
    print(f"  Legs filled        : {n_filled} of 2")
    print(f"  Long  leg          : {fmt_leg(long_leg)}")
    print(f"  Short leg          : {fmt_leg(short_leg)}")
    print(f"  Day P&L            : ${day_usd:>+12,.2f}  (${day_oz:+.2f}/oz on {LOT_SIZE} lot)")
    if balance_state is not None:
        end_bal = balance_state["current_balance"]
        cum = round(end_bal - balance_state["starting_capital"], 2)
        cum_pct = (cum / balance_state["starting_capital"] * 100.0) if balance_state["starting_capital"] else 0.0
        print(f"  Ending balance     : ${end_bal:>+12,.2f}")
        print(f"  Lifetime cum P&L   : ${cum:>+12,.2f}   ({cum_pct:+.2f}% on starting capital "
              f"${balance_state['starting_capital']:,.2f})")
        print(f"  Lifetime trades    : {balance_state.get('trade_count', 0)}")
    print(f"  Commission         : ${COMMISSION_PER_TRADE:.2f}/oz on each closed leg "
          f"(spread captured in bid/ask fills)")
    print("=" * 72)


def daily_summary_dict(long_leg: dict, short_leg: dict,
                       day_start_balance: float = None,
                       balance_state: dict = None) -> dict:
    """Same data as print_daily_summary, but as a dict for event logging."""
    day_oz = (long_leg.get("net_pnl_oz", 0) or 0) + (short_leg.get("net_pnl_oz", 0) or 0)
    day_usd = (long_leg.get("net_pnl_usd", 0) or 0) + (short_leg.get("net_pnl_usd", 0) or 0)

    def leg_dict(l):
        return {
            "status": l["status"],
            "entry_price": l.get("entry_price"),
            "entry_bid": l.get("entry_bid"),
            "entry_ask": l.get("entry_ask"),
            "exit_price": l.get("exit_price"),
            "exit_bid": l.get("exit_bid"),
            "exit_ask": l.get("exit_ask"),
            "exit_reason": l.get("exit_reason"),
            "gross_pnl_oz": l.get("gross_pnl_oz"),
            "net_pnl_oz": l.get("net_pnl_oz"),
            "net_pnl_usd": l.get("net_pnl_usd"),
            "lock_armed": l.get("lock_armed", False),
            "extreme_seen": l.get("extreme_price"),
            "mfe_price": l.get("mfe_price"),
            "mae_price": l.get("mae_price"),
        }

    out = {
        "anchor_price": long_leg["anchor_price"],
        "buy_stop_price": long_leg["trigger_price"],
        "sell_stop_price": short_leg["trigger_price"],
        "long": leg_dict(long_leg),
        "short": leg_dict(short_leg),
        "day_pnl_oz": round(day_oz, 2),
        "day_pnl_usd": round(day_usd, 2),
        "legs_filled": sum(1 for l in (long_leg, short_leg) if l["status"] == "closed"),
    }
    if balance_state is not None:
        out["day_start_balance"] = round(day_start_balance, 2) if day_start_balance is not None else None
        out["day_end_balance"] = round(balance_state["current_balance"], 2)
        out["cumulative_pnl_usd"] = round(
            balance_state["current_balance"] - balance_state["starting_capital"], 2
        )
        out["starting_capital"] = balance_state["starting_capital"]
        out["trade_count_lifetime"] = balance_state.get("trade_count", 0)
    return out


# ============================================================================
# HEARTBEAT (live one-liner)
# ============================================================================

def print_heartbeat(bid: float, ask: float, anchor_price: float,
                    long_leg: dict, short_leg: dict,
                    contract_size: float,
                    balance_state: dict = None):
    """Single line, overwrites in place, shows both legs and the spread.
    Balance is appended at the end when balance_state is supplied."""
    ist = get_ist_now()
    spread = ask - bid
    delta_anc = bid - anchor_price

    def leg_block(leg: dict) -> str:
        s = leg["status"]
        if s == "pending":
            if leg["side"] == "long":
                dist = leg["trigger_price"] - ask  # need ask to reach trigger
                return f"L pend (need +${dist:.2f})"
            else:
                dist = bid - leg["trigger_price"]  # need bid to reach trigger
                return f"S pend (need -${dist:.2f})"
        if s == "active":
            if leg["side"] == "long":
                unr = bid - leg["entry_price"]  # mark-to-bid for long
            else:
                unr = leg["entry_price"] - ask  # mark-to-ask for short
            lock_str = f"lk${leg['lock_price']:.2f}" if leg["lock_armed"] else "lk—"
            return (
                f"{'L' if leg['side']=='long' else 'S'} act "
                f"@${leg['entry_price']:.2f} ext${leg['extreme_price']:.2f} "
                f"{lock_str} unr${unr:+.2f}"
            )
        if s == "closed":
            return (
                f"{'L' if leg['side']=='long' else 'S'} closed "
                f"${leg.get('net_pnl_oz', 0):+.2f} ({leg.get('exit_reason', '?')})"
            )
        if s == "cancelled":
            return f"{'L' if leg['side']=='long' else 'S'} cancelled"
        return f"{'L' if leg['side']=='long' else 'S'} ?"

    day_oz = (long_leg.get("net_pnl_oz", 0) or 0) + (short_leg.get("net_pnl_oz", 0) or 0)

    bal_part = ""
    if balance_state is not None:
        bal_part = f"  bal ${balance_state['current_balance']:,.2f}"

    line = (
        f"\r[{ist:%H:%M:%S}] b${bid:.2f}/a${ask:.2f} sp${spread:.2f}  "
        f"anc ${anchor_price:.2f} Δ{delta_anc:+.2f}  "
        f"{leg_block(long_leg)} | {leg_block(short_leg)}  "
        f"day ${day_oz:+.2f}/oz"
        f"{bal_part}"
    )
    sys.stdout.write(line.ljust(HEARTBEAT_WIDTH))
    sys.stdout.flush()


# ============================================================================
# BACKFILL  (mid-day restart recovery)
# ============================================================================

def backfill_from_bars(rates, anchor_price: float, anchor_time: datetime,
                       contract_size: float):
    """Replay M1 bars from anchor -> now to recover state after a mid-day
    restart. APPROXIMATE: M1 bars don't carry bid/ask; we treat the bar close
    as bid and apply FIXED_SPREAD_FALLBACK to derive ask. Live-stream trades
    are precise; backfill trades are tagged backfill=True so you can separate
    them in analysis."""
    buy_stop = round(anchor_price + TRIGGER_DISTANCE, 2)
    sell_stop = round(anchor_price - TRIGGER_DISTANCE, 2)

    long_leg = make_leg("long", buy_stop, anchor_price)
    short_leg = make_leg("short", sell_stop, anchor_price)
    long_leg["backfill"] = True
    short_leg["backfill"] = True

    trade_id_counter = 1
    closed_legs = []

    if rates is None or len(rates) == 0:
        return long_leg, short_leg, closed_legs, trade_id_counter, None

    last_close = anchor_price
    sp = FIXED_SPREAD_FALLBACK

    for bar in rates:
        bar_time = broker_dt(int(bar["time"]))
        if bar_time < anchor_time:
            continue
        bar_ist = bar_time.replace(tzinfo=timezone.utc).astimezone(IST)
        o = float(bar["open"])
        h = float(bar["high"])
        l = float(bar["low"])
        c = float(bar["close"])
        last_close = c

        # In backfill, treat bar low as the lowest bid seen and bar high +
        # spread as the highest ask seen during the bar.
        bar_bid_lo, bar_bid_hi = l, h
        bar_ask_lo, bar_ask_hi = l + sp, h + sp

        # ---- ENTRIES (using bar ask for long trigger, bar bid for short) ----
        buy_could  = long_leg["status"] == "pending"  and bar_ask_hi >= buy_stop
        sell_could = short_leg["status"] == "pending" and bar_bid_lo <= sell_stop

        if buy_could and sell_could:
            buy_first = (o >= anchor_price)
        elif buy_could:
            buy_first = True
        elif sell_could:
            buy_first = False
        else:
            buy_first = None

        def _fill_long():
            long_leg["status"] = "active"
            long_leg["entry_bid"] = round(buy_stop - sp, 2)
            long_leg["entry_ask"] = buy_stop
            long_leg["spread_at_entry"] = sp
            long_leg["entry_price"] = buy_stop
            long_leg["entry_server_time"] = bar_time
            long_leg["entry_ist_time"] = bar_ist
            long_leg["extreme_price"] = max(buy_stop, h)
            long_leg["trade_id"] = trade_id_counter

        def _fill_short():
            short_leg["status"] = "active"
            short_leg["entry_bid"] = sell_stop
            short_leg["entry_ask"] = round(sell_stop + sp, 2)
            short_leg["spread_at_entry"] = sp
            short_leg["entry_price"] = sell_stop
            short_leg["entry_server_time"] = bar_time
            short_leg["entry_ist_time"] = bar_ist
            short_leg["extreme_price"] = min(sell_stop, l + sp)
            short_leg["trade_id"] = trade_id_counter

        if buy_first is True and buy_could:
            _fill_long(); trade_id_counter += 1
        elif buy_first is False and sell_could:
            _fill_short(); trade_id_counter += 1

        if buy_could and long_leg["status"] == "pending":
            _fill_long(); trade_id_counter += 1
        if sell_could and short_leg["status"] == "pending":
            _fill_short(); trade_id_counter += 1

        # ---- LOCK UPDATES against bar extremes ----
        for leg in (long_leg, short_leg):
            if leg["status"] != "active":
                continue
            if leg["side"] == "long":
                if h > leg["extreme_price"]:
                    leg["extreme_price"] = h
                profit = leg["extreme_price"] - leg["entry_price"]
                if not leg["lock_armed"]:
                    if profit >= TRAIL_ACTIVATION:
                        leg["lock_armed"] = True
                        leg["lock_price"] = round(leg["extreme_price"] - TRAIL_DISTANCE, 2)
                else:
                    new_lock = round(leg["extreme_price"] - TRAIL_DISTANCE, 2)
                    if new_lock > leg["lock_price"]:
                        leg["lock_price"] = new_lock
            else:
                # Short extreme tracked on ask = bar_low + spread
                if (l + sp) < leg["extreme_price"]:
                    leg["extreme_price"] = l + sp
                profit = leg["entry_price"] - leg["extreme_price"]
                if not leg["lock_armed"]:
                    if profit >= TRAIL_ACTIVATION:
                        leg["lock_armed"] = True
                        leg["lock_price"] = round(leg["extreme_price"] + TRAIL_DISTANCE, 2)
                else:
                    new_lock = round(leg["extreme_price"] + TRAIL_DISTANCE, 2)
                    if new_lock < leg["lock_price"]:
                        leg["lock_price"] = new_lock

        # ---- EXITS (lock before hard SL); approximate fill = trigger level ----
        for leg in (long_leg, short_leg):
            if leg["status"] != "active":
                continue
            if leg["side"] == "long":
                if leg["lock_armed"] and bar_bid_lo <= leg["lock_price"]:
                    opp = short_leg["status"] in ("active", "closed")
                    close_leg(leg, leg["lock_price"], leg["lock_price"] + sp,
                              "lock", bar_time, bar_ist, contract_size, opp)
                    closed_legs.append(leg)
                elif bar_bid_lo <= leg["hard_sl_price"]:
                    opp = short_leg["status"] in ("active", "closed")
                    close_leg(leg, leg["hard_sl_price"],
                              leg["hard_sl_price"] + sp,
                              "hard_sl", bar_time, bar_ist, contract_size, opp)
                    closed_legs.append(leg)
            else:
                if leg["lock_armed"] and bar_ask_hi >= leg["lock_price"]:
                    opp = long_leg["status"] in ("active", "closed")
                    close_leg(leg, leg["lock_price"] - sp, leg["lock_price"],
                              "lock", bar_time, bar_ist, contract_size, opp)
                    closed_legs.append(leg)
                elif bar_ask_hi >= leg["hard_sl_price"]:
                    opp = long_leg["status"] in ("active", "closed")
                    close_leg(leg, leg["hard_sl_price"] - sp,
                              leg["hard_sl_price"],
                              "hard_sl", bar_time, bar_ist, contract_size, opp)
                    closed_legs.append(leg)

    return long_leg, short_leg, closed_legs, trade_id_counter, last_close


# ============================================================================
# MAIN LOOP
# ============================================================================

def run():
    connect_mt5()
    ensure_trade_log_header(TRADE_LOG_PATH)

    info = mt5.symbol_info(SYMBOL)
    contract_size = float(info.trade_contract_size) if info else 100.0

    daily_records = load_daily_log(LOG_PATH)
    recent_events = load_recent_events(EVENT_LOG_PATH)
    continuity = verify_continuity(recent_events)
    balance_state = load_balance()
    # Daily XLSX rows persist across restarts by being rebuilt from the CSV.
    # Whenever a day finalizes we append the fresh row and rewrite the XLSX.
    daily_xlsx_rows = load_xlsx_rows_from_csv()

    active_anchor = None
    anchor_price = None
    buy_stop_price = None
    sell_stop_price = None
    long_leg = None
    short_leg = None
    trade_id_counter = 1
    eod_closed_today = False
    last_bid = None
    last_ask = None
    day_start_balance = balance_state["current_balance"]  # snapshot at session start
    mt5_balance_at_start = None  # set per-session from real MT5 balance

    # === Startup banner ===
    starting_cap = balance_state["starting_capital"]
    current_bal = balance_state["current_balance"]
    cum_pnl = current_bal - starting_cap
    mt5_bal_now = get_mt5_balance()
    mt5_eq_now = get_mt5_equity()
    print("=" * 72)
    print(f"  AUREON MODIFIED  ·  PAPER MOCK  ·  {SYMBOL}")
    print("=" * 72)
    print(f"  Starting capital: ${starting_cap:>12,.2f}  (first-init only — balance file is source of truth)")
    print(f"  Paper balance   : ${current_bal:>+12,.2f}  (cumulative P&L ${cum_pnl:>+,.2f}, "
          f"{balance_state.get('trade_count', 0)} closed trades to date)")
    if mt5_bal_now is not None:
        print(f"  MT5 balance     : ${mt5_bal_now:>+12,.2f}  (live account — paper will re-sync to this each 02:00)")
    else:
        print(f"  MT5 balance     : unavailable (account_info returned None — paper-only)")
    if mt5_eq_now is not None:
        print(f"  MT5 equity      : ${mt5_eq_now:>+12,.2f}  (balance + floating P&L)")
    print(f"  Anchor time     : {ANCHOR_HOUR:02d}:{ANCHOR_MINUTE:02d} broker (M1 open)")
    print(f"  EOD time        : {EOD_HOUR:02d}:{EOD_MINUTE:02d} broker  (force-close + cancel pending)")
    print(f"  Trigger / SL    : ±${TRIGGER_DISTANCE:.2f} / ±${HARD_SL:.2f}")
    print(f"  Trail           : ${TRAIL_DISTANCE:.2f} after ${TRAIL_ACTIVATION:.2f} profit")
    print(f"  Cost model      : bid/ask spread modelled live  +  commission ${COMMISSION_PER_TRADE:.2f}/oz")
    print(f"  Fallback spread : ${FIXED_SPREAD_FALLBACK:.2f}  (when broker quote feed reports ask<=bid)")
    print(f"  Lot size        : {LOT_SIZE}  (contract {contract_size:g} → ${LOT_SIZE*contract_size:.0f}/$1 move)")
    print(f"  Mode            : OCO OFF (modified) — both stops independent")
    print(f"  Trades          : SIMULATED ONLY — NO REAL ORDERS PLACED")
    print(f"  Files           : {LOG_PATH}")
    print(f"                    {TRADE_LOG_PATH}")
    print(f"                    {EVENT_LOG_PATH}")
    print(f"  Continuity      : {continuity}")
    print("=" * 72)

    log_event("startup", {
        "symbol": SYMBOL,
        "lot_size": LOT_SIZE,
        "contract_size": contract_size,
        "starting_capital": balance_state["starting_capital"],
        "current_balance_at_startup": balance_state["current_balance"],
        "cumulative_pnl_at_startup": round(
            balance_state["current_balance"] - balance_state["starting_capital"], 2),
        "lifetime_trades_at_startup": balance_state.get("trade_count", 0),
        "config": {
            "trigger_distance": TRIGGER_DISTANCE,
            "hard_sl": HARD_SL,
            "trail_distance": TRAIL_DISTANCE,
            "trail_activation": TRAIL_ACTIVATION,
            "commission_per_trade": COMMISSION_PER_TRADE,
            "fixed_spread_fallback": FIXED_SPREAD_FALLBACK,
            "anchor_hour": ANCHOR_HOUR,
            "eod_hour": EOD_HOUR,
            "eod_minute": EOD_MINUTE,
        },
        "continuity_note": continuity,
        "backfill_on_start": BACKFILL_ON_START,
    }, datetime.now(timezone.utc).replace(tzinfo=None), get_ist_now())

    while True:
        try:
            server_now = get_server_now(SYMBOL)
            anchor_time = get_current_anchor_time(server_now)
            ist_now = get_ist_now()

            # ================================================================
            # DAILY ROLLOVER (02:00)
            # ================================================================
            if active_anchor != anchor_time:
                # ---- Finalize previous day (if any) ----
                if active_anchor is not None:
                    closing_bid = last_bid if last_bid is not None else anchor_price
                    closing_ask = last_ask if last_ask is not None else (closing_bid + FIXED_SPREAD_FALLBACK)
                    # Defensive close: if 23:00 EOD missed for some reason,
                    # close any still-active legs at the latest known prices.
                    if long_leg and long_leg["status"] == "active":
                        opp = short_leg["status"] in ("active", "closed")
                        close_leg(long_leg, closing_bid, closing_ask,
                                  "eod_long_rollover", server_now, ist_now,
                                  contract_size, opp)
                        append_trade_log(long_leg, active_anchor.date().isoformat())
                        book_trade(long_leg, balance_state)
                    if short_leg and short_leg["status"] == "active":
                        opp = long_leg["status"] in ("active", "closed")
                        close_leg(short_leg, closing_bid, closing_ask,
                                  "eod_short_rollover", server_now, ist_now,
                                  contract_size, opp)
                        append_trade_log(short_leg, active_anchor.date().isoformat())
                        book_trade(short_leg, balance_state)

                    for leg in (long_leg, short_leg):
                        if leg and leg["status"] == "pending":
                            leg["status"] = "cancelled"

                    prev_date = active_anchor.date().isoformat()
                    print_daily_summary(prev_date, long_leg, short_leg, contract_size,
                                        day_start_balance=day_start_balance,
                                        balance_state=balance_state)
                    log_event("day_finalized", {
                        "session_date": prev_date,
                        **daily_summary_dict(long_leg, short_leg,
                                             day_start_balance, balance_state),
                    }, server_now, ist_now)

                    record = build_daily_record(active_anchor, anchor_price,
                                                 buy_stop_price, sell_stop_price,
                                                 long_leg, short_leg,
                                                 closing_bid, server_now, ist_now,
                                                 day_start_balance, balance_state)
                    daily_records[record["date"]] = record
                    save_daily_log(LOG_PATH, daily_records)

                    # XLSX update for the rollover-closed day too. notes="" if
                    # EOD already wrote it; this is the safety-net path.
                    xlsx_row = daily_xlsx_row(
                        active_anchor, anchor_price, buy_stop_price, sell_stop_price,
                        long_leg, short_leg, day_start_balance, balance_state,
                        contract_size,
                        mt5_balance_at_start=mt5_balance_at_start,
                        mt5_equity_at_close=get_mt5_equity(),
                        notes="rollover close (defensive)",
                    )
                    prev_date = active_anchor.date().isoformat()
                    daily_xlsx_rows = [r for r in daily_xlsx_rows if r.get("Date") != prev_date]
                    daily_xlsx_rows.append(xlsx_row)
                    update_daily_xlsx(daily_xlsx_rows, balance_state, contract_size)

                # ---- Start new day ----
                anchor_data = fetch_anchor(SYMBOL, anchor_time)
                active_anchor = anchor_time
                anchor_price = anchor_data["price"]
                buy_stop_price = round(anchor_price + TRIGGER_DISTANCE, 2)
                sell_stop_price = round(anchor_price - TRIGGER_DISTANCE, 2)
                long_leg = make_leg("long", buy_stop_price, anchor_price)
                short_leg = make_leg("short", sell_stop_price, anchor_price)
                eod_closed_today = False
                trade_id_counter = 1
                # === Sync day_start_balance to the real MT5 account ===
                # The paper bot's "current_balance" is the running paper P&L
                # applied on top of the most recent MT5 reset. Each new
                # session, re-anchor to whatever the real account currently
                # has — this way the paper trajectory tracks the real one,
                # and any deposits/withdrawals/manual trades that moved the
                # MT5 balance get picked up automatically.
                mt5_balance = get_mt5_balance()
                if mt5_balance is not None:
                    if abs(balance_state["current_balance"] - mt5_balance) > 0.01:
                        print(f"    syncing balance to MT5: "
                              f"${balance_state['current_balance']:,.2f} → ${mt5_balance:,.2f}")
                    balance_state["current_balance"] = mt5_balance
                    save_balance(balance_state)
                day_start_balance = balance_state["current_balance"]
                # Stash for the XLSX row at EOD; equity at close is captured later.
                mt5_balance_at_start = mt5_balance

                print()
                print(f">>> NEW SESSION  {active_anchor.date().isoformat()}  "
                      f"{fmt_times(active_anchor, ist_now)}")
                print(f"    anchor ${anchor_price:.2f} @ {active_anchor}")
                print(f"    buy stop  ${buy_stop_price:.2f}    sell stop ${sell_stop_price:.2f}")
                print(f"    HARD SL   long ${long_leg['hard_sl_price']:.2f}  "
                      f"short ${short_leg['hard_sl_price']:.2f}")
                print(f"    day-start balance ${day_start_balance:,.2f}")

                log_event("session_start", {
                    "session_date": active_anchor.date().isoformat(),
                    "anchor_server_time": active_anchor.isoformat(sep=" "),
                    "anchor_price": anchor_price,
                    "buy_stop_price": buy_stop_price,
                    "sell_stop_price": sell_stop_price,
                    "lot_size": LOT_SIZE,
                }, server_now, ist_now)

                # ---- One-shot backfill if booting mid-day ----
                if BACKFILL_ON_START and server_now > anchor_time + timedelta(minutes=2):
                    rates = mt5.copy_rates_range(
                        SYMBOL, TIMEFRAME,
                        to_mt5_dt(active_anchor), to_mt5_dt(server_now),
                    )
                    if rates is not None and len(rates) > 0:
                        bf_long, bf_short, bf_closed, bf_tid, bf_last = backfill_from_bars(
                            rates, anchor_price, active_anchor, contract_size,
                        )
                        long_leg = bf_long
                        short_leg = bf_short
                        trade_id_counter = bf_tid
                        if bf_last is not None:
                            last_bid = bf_last
                            last_ask = bf_last + FIXED_SPREAD_FALLBACK

                        print(f">>> BACKFILL  replayed {len(rates)} bars  "
                              f"{len(bf_closed)} closed leg(s)")
                        for leg in bf_closed:
                            append_trade_log(leg, active_anchor.date().isoformat())
                            book_trade(leg, balance_state)
                            print(
                                f"    backfill {leg['side']:<5} entry ${leg['entry_price']:.2f} "
                                f"-> exit ${leg['exit_price']:.2f} ({leg['exit_reason']})  "
                                f"net ${leg.get('net_pnl_oz', 0):+.2f}/oz"
                            )
                            log_event("leg_exit", {
                                "trade_id": leg["trade_id"],
                                "session_date": active_anchor.date().isoformat(),
                                "leg": leg["side"],
                                "entry_price": leg["entry_price"],
                                "exit_price": leg["exit_price"],
                                "exit_reason": leg["exit_reason"],
                                "gross_pnl_oz": leg["gross_pnl_oz"],
                                "net_pnl_oz": leg["net_pnl_oz"],
                                "net_pnl_usd": leg["net_pnl_usd"],
                                "backfill": True,
                            }, leg["exit_server_time"], leg["exit_ist_time"])
                        for leg in (long_leg, short_leg):
                            if leg["status"] == "active":
                                print(f"    backfill {leg['side']:<5} STILL OPEN  "
                                      f"@${leg['entry_price']:.2f}  "
                                      f"ext${leg['extreme_price']:.2f}  "
                                      f"lock{'$%.2f' % leg['lock_price'] if leg['lock_armed'] else '—'}")

            # ================================================================
            # PER-TICK PROCESSING
            # ================================================================
            bid, ask = get_bid_ask(SYMBOL)
            if bid is None:
                time.sleep(POLL_SECONDS)
                continue
            last_bid = bid
            last_ask = ask

            # ---- EOD trigger (23:00 broker, fires once per day) ----
            in_eod_window = (server_now.hour >= EOD_HOUR
                             and server_now.minute >= EOD_MINUTE)
            # Guard: only fire if we're past EOD_HOUR:EOD_MINUTE *today* and
            # before the 02:00 of the next day. server_now.hour >= 23 already
            # excludes pre-23h. Don't fire again after midnight (next 02:00
            # rollover resets the flag via active_anchor change).
            if in_eod_window and not eod_closed_today:
                print()
                print(f">>> EOD CLOSE WINDOW REACHED  "
                      f"{fmt_times(server_now, ist_now)}")
                for leg in (long_leg, short_leg):
                    if leg["status"] == "active":
                        opp = (short_leg if leg is long_leg else long_leg)["status"] in ("active", "closed")
                        reason = "eod_long" if leg["side"] == "long" else "eod_short"
                        close_leg(leg, bid, ask, reason, server_now, ist_now,
                                  contract_size, opp)
                        append_trade_log(leg, active_anchor.date().isoformat())
                        book_trade(leg, balance_state)
                        print(f"    {leg['side']:<5}  closed @${leg['exit_price']:.2f}  "
                              f"net ${leg['net_pnl_oz']:+.2f}/oz "
                              f"(${leg['net_pnl_usd']:+.2f})  "
                              f"balance now ${balance_state['current_balance']:,.2f}")
                        log_event("leg_exit", {
                            "trade_id": leg["trade_id"],
                            "session_date": active_anchor.date().isoformat(),
                            "leg": leg["side"],
                            "entry_price": leg["entry_price"],
                            "exit_price": leg["exit_price"],
                            "exit_bid": leg["exit_bid"],
                            "exit_ask": leg["exit_ask"],
                            "spread_at_exit": leg["spread_at_exit"],
                            "exit_reason": leg["exit_reason"],
                            "gross_pnl_oz": leg["gross_pnl_oz"],
                            "net_pnl_oz": leg["net_pnl_oz"],
                            "net_pnl_usd": leg["net_pnl_usd"],
                            "mfe_price": leg["mfe_price"],
                            "mae_price": leg["mae_price"],
                            "duration_seconds": leg["duration_seconds"],
                        }, server_now, ist_now)
                    elif leg["status"] == "pending":
                        leg["status"] = "cancelled"
                        log_event("leg_cancelled", {
                            "session_date": active_anchor.date().isoformat(),
                            "leg": leg["side"],
                            "reason": "eod_no_trigger",
                        }, server_now, ist_now)
                eod_closed_today = True
                print_daily_summary(active_anchor.date().isoformat(),
                                    long_leg, short_leg, contract_size,
                                    day_start_balance=day_start_balance,
                                    balance_state=balance_state)
                log_event("eod_reached", {
                    "session_date": active_anchor.date().isoformat(),
                    **daily_summary_dict(long_leg, short_leg,
                                         day_start_balance, balance_state),
                }, server_now, ist_now)

                # --- Append finalized day to XLSX and save ---
                mt5_equity_now = get_mt5_equity()
                xlsx_row = daily_xlsx_row(
                    active_anchor, anchor_price, buy_stop_price, sell_stop_price,
                    long_leg, short_leg, day_start_balance, balance_state,
                    contract_size,
                    mt5_balance_at_start=mt5_balance_at_start,
                    mt5_equity_at_close=mt5_equity_now,
                    notes="EOD 23:00 close",
                )
                # Replace any existing row for this date (idempotent on retry)
                today_str = active_anchor.date().isoformat()
                daily_xlsx_rows = [r for r in daily_xlsx_rows if r.get("Date") != today_str]
                daily_xlsx_rows.append(xlsx_row)
                update_daily_xlsx(daily_xlsx_rows, balance_state, contract_size)

            # ---- Per-leg state transitions (skipped after EOD) ----
            if not eod_closed_today:
                for leg in (long_leg, short_leg):
                    if leg["status"] == "pending":
                        if trigger_hit(leg, bid, ask):
                            opp = (short_leg if leg is long_leg else long_leg)
                            # Realistic entry fill: long fills at ask, short at bid.
                            fill_price = ask if leg["side"] == "long" else bid
                            leg["status"] = "active"
                            leg["entry_bid"] = round(bid, 2)
                            leg["entry_ask"] = round(ask, 2)
                            leg["spread_at_entry"] = round(ask - bid, 2)
                            leg["entry_price"] = round(fill_price, 2)
                            leg["entry_server_time"] = server_now
                            leg["entry_ist_time"] = ist_now
                            leg["extreme_price"] = (bid if leg["side"] == "long" else ask)
                            leg["trade_id"] = trade_id_counter
                            trade_id_counter += 1
                            print()
                            print(f">>> ENTRY  {leg['side']:<5}  "
                                  f"@${leg['entry_price']:.2f}  "
                                  f"(bid ${bid:.2f}, ask ${ask:.2f}, sp ${ask-bid:.2f})  "
                                  f"anchor ${anchor_price:.2f}  "
                                  f"hard SL ${leg['hard_sl_price']:.2f}  "
                                  f"{fmt_times(server_now, ist_now)}")
                            log_event("leg_triggered", {
                                "trade_id": leg["trade_id"],
                                "session_date": active_anchor.date().isoformat(),
                                "leg": leg["side"],
                                "entry_price": leg["entry_price"],
                                "entry_bid": leg["entry_bid"],
                                "entry_ask": leg["entry_ask"],
                                "spread_at_entry": leg["spread_at_entry"],
                                "anchor_price": anchor_price,
                                "hard_sl_price": leg["hard_sl_price"],
                                "opposite_already_fired": opp["status"] in ("active", "closed"),
                            }, server_now, ist_now)

                    elif leg["status"] == "active":
                        armed_now, advanced_now = update_extreme_and_lock(leg, bid, ask)
                        if armed_now:
                            print()
                            print(f">>> LOCK ARMED  {leg['side']:<5}  "
                                  f"ext ${leg['extreme_price']:.2f}  "
                                  f"lock ${leg['lock_price']:.2f}  "
                                  f"{fmt_times(server_now, ist_now)}")
                            log_event("leg_lock_armed", {
                                "trade_id": leg["trade_id"],
                                "session_date": active_anchor.date().isoformat(),
                                "leg": leg["side"],
                                "extreme_price": leg["extreme_price"],
                                "lock_price": leg["lock_price"],
                                "bid": bid, "ask": ask,
                            }, server_now, ist_now)

                        if lock_hit(leg, bid, ask):
                            opp = (short_leg if leg is long_leg else long_leg)
                            opp_fired = opp["status"] in ("active", "closed")
                            close_leg(leg, bid, ask, "lock",
                                      server_now, ist_now, contract_size, opp_fired)
                            append_trade_log(leg, active_anchor.date().isoformat())
                            book_trade(leg, balance_state)
                            print()
                            print(f">>> EXIT   {leg['side']:<5}  "
                                  f"@${leg['exit_price']:.2f}  reason LOCK  "
                                  f"(bid ${bid:.2f}/ask ${ask:.2f})  "
                                  f"net ${leg['net_pnl_oz']:+.2f}/oz "
                                  f"(${leg['net_pnl_usd']:+.2f})  "
                                  f"bal ${balance_state['current_balance']:,.2f}  "
                                  f"{fmt_times(server_now, ist_now)}")
                            log_event("leg_exit", {
                                "trade_id": leg["trade_id"],
                                "session_date": active_anchor.date().isoformat(),
                                "leg": leg["side"],
                                "entry_price": leg["entry_price"],
                                "exit_price": leg["exit_price"],
                                "exit_bid": leg["exit_bid"],
                                "exit_ask": leg["exit_ask"],
                                "spread_at_exit": leg["spread_at_exit"],
                                "exit_reason": "lock",
                                "gross_pnl_oz": leg["gross_pnl_oz"],
                                "net_pnl_oz": leg["net_pnl_oz"],
                                "net_pnl_usd": leg["net_pnl_usd"],
                                "mfe_price": leg["mfe_price"],
                                "mae_price": leg["mae_price"],
                                "duration_seconds": leg["duration_seconds"],
                                "extreme_seen": leg["extreme_price"],
                                "opposite_leg_fired": opp_fired,
                            }, server_now, ist_now)
                        elif hard_sl_hit(leg, bid, ask):
                            opp = (short_leg if leg is long_leg else long_leg)
                            opp_fired = opp["status"] in ("active", "closed")
                            close_leg(leg, bid, ask, "hard_sl",
                                      server_now, ist_now, contract_size, opp_fired)
                            append_trade_log(leg, active_anchor.date().isoformat())
                            book_trade(leg, balance_state)
                            print()
                            print(f">>> EXIT   {leg['side']:<5}  "
                                  f"@${leg['exit_price']:.2f}  reason HARD_SL  "
                                  f"(bid ${bid:.2f}/ask ${ask:.2f})  "
                                  f"net ${leg['net_pnl_oz']:+.2f}/oz "
                                  f"(${leg['net_pnl_usd']:+.2f})  "
                                  f"bal ${balance_state['current_balance']:,.2f}  "
                                  f"{fmt_times(server_now, ist_now)}")
                            log_event("leg_exit", {
                                "trade_id": leg["trade_id"],
                                "session_date": active_anchor.date().isoformat(),
                                "leg": leg["side"],
                                "entry_price": leg["entry_price"],
                                "exit_price": leg["exit_price"],
                                "exit_bid": leg["exit_bid"],
                                "exit_ask": leg["exit_ask"],
                                "spread_at_exit": leg["spread_at_exit"],
                                "exit_reason": "hard_sl",
                                "gross_pnl_oz": leg["gross_pnl_oz"],
                                "net_pnl_oz": leg["net_pnl_oz"],
                                "net_pnl_usd": leg["net_pnl_usd"],
                                "mfe_price": leg["mfe_price"],
                                "mae_price": leg["mae_price"],
                                "duration_seconds": leg["duration_seconds"],
                                "extreme_seen": leg["extreme_price"],
                                "opposite_leg_fired": opp_fired,
                            }, server_now, ist_now)

            # ---- Persist running daily record ----
            # Use mid-price for the "last_price" column (purely cosmetic).
            mid = round((bid + ask) / 2, 2)
            record = build_daily_record(active_anchor, anchor_price,
                                         buy_stop_price, sell_stop_price,
                                         long_leg, short_leg,
                                         mid, server_now, ist_now,
                                         day_start_balance, balance_state)
            daily_records[record["date"]] = record
            save_daily_log(LOG_PATH, daily_records)

            print_heartbeat(bid, ask, anchor_price, long_leg, short_leg,
                            contract_size, balance_state)
            time.sleep(POLL_SECONDS)

        except KeyboardInterrupt:
            today = active_anchor.date().isoformat() if active_anchor else "(no session)"
            if long_leg is not None and short_leg is not None:
                print_daily_summary(today, long_leg, short_leg, contract_size,
                                    day_start_balance=day_start_balance,
                                    balance_state=balance_state)
                log_event("shutdown", {
                    "reason": "keyboard_interrupt",
                    "session_date": today,
                    **daily_summary_dict(long_leg, short_leg,
                                         day_start_balance, balance_state),
                }, datetime.utcnow(), get_ist_now())
            raise
        except Exception as e:
            print(f"\n[{get_ist_now():%H:%M:%S}] tracker error: {e}")
            log_event("error", {
                "message": str(e),
                "type": type(e).__name__,
            }, datetime.utcnow(), get_ist_now())
            time.sleep(RETRY_SECONDS)


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print("\nStopped manually.")
    finally:
        mt5.shutdown()